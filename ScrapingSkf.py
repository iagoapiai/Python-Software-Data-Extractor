import pyautogui
import openpyxl
import time
import pyperclip

Ancora_rolamento = None
Num_Rolamento = None

linha = 2
rolamento = 0
modelo = 0
anchor = 0

fab_coletados = []

workbook = openpyxl.Workbook()
sheet = workbook.active

sheet["A1"] = 'Fabricante'
sheet["B1"] = 'N° Rolamento'        
sheet["C1"] = 'BPFO'
sheet["D1"] = 'BPFI'
sheet["E1"] = 'BSF'
sheet["F1"] = 'FTF'

def criar_workbook():
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet["A1"] = 'Fabricante'
    sheet["B1"] = 'N° Rolamento'        
    sheet["C1"] = 'BPFO'
    sheet["D1"] = 'BPFI'
    sheet["E1"] = 'BSF'
    sheet["F1"] = 'FTF'

#  Clicando no rolamento na página inicial
while Num_Rolamento != "U-4034-A.1":
    
    if anchor == 1:
        pyautogui.click(934,371)                 
        time.sleep(0.5)                                                                                                                       
        pyautogui.press('down')
        time.sleep(4)
        anchor = 0 
        pyautogui.click(1083,360) 
    
    print(f"Coletando {rolamento}° rolamento!")
    
    if 0 <= rolamento <= 7:
        pyautogui.click(1027,441)
        time.sleep(0.3)
        if rolamento != 0:
            for _ in range(rolamento):
                pyautogui.press('down')
    
    else:
        pyautogui.click(992,534)
        time.sleep(0.1)
        pyautogui.press('down')
        
    # Clicar em adicionar
    pyautogui.press('TAB')
    pyautogui.press('TAB')
    pyautogui.press('ENTER')
    time.sleep(0.30)

    # Copiando Nome do Rolamento
    pyautogui.hotkey('ctrl', 'c')
    time.sleep(0.1)
    Num_Rolamento = pyperclip.paste()
    print(Num_Rolamento)
    time.sleep(0.1)
    pyautogui.press('TAB')
    
    # Copiando Fabricante
    pyautogui.hotkey('ctrl', 'c')
    time.sleep(0.1)
    Fabricante = pyperclip.paste()
    print(Fabricante)
    time.sleep(0.1)
        
    c = 0
    if Num_Rolamento == Ancora_rolamento:
        workbook.save(f"{Fabricante}.xlsx")
        linha = 2
        c = 1 
        anchor = 1
        rolamento = 0
        pyautogui.click(1030,690)
        criar_workbook()
        time.sleep(1)
        
    if Fabricante in fab_coletados:
        anchor = 1
        c = 1 
    
    if c == 0:      
        Ancora_rolamento = Num_Rolamento
        
        # Copiando BPFO
        pyautogui.click(926,616)
        pyautogui.click(926,616)
        pyautogui.click(button='RIGHT')
        time.sleep(0.2)
        for _ in range(3):
            pyautogui.press('down')
        pyautogui.press('ENTER')
        BPFO = pyperclip.paste()
        time.sleep(0.1)
        print(BPFO)

        # Copiando BPFI
        pyautogui.click(919,641)
        pyautogui.click(919,641)
        pyautogui.click(button='RIGHT')
        time.sleep(0.2)
        for _ in range(3):
            pyautogui.press('down')
        pyautogui.press('ENTER')
        BPFI = pyperclip.paste()
        time.sleep(0.1)
        print(BPFI)

        # Copiando BSF
        pyautogui.click(1089,620)
        pyautogui.click(1089,620)
        pyautogui.click(button='RIGHT')
        time.sleep(0.2)
        for _ in range(3):
            pyautogui.press('down')
        pyautogui.press('ENTER')
        BSF = pyperclip.paste()
        time.sleep(0.1)
        print(BSF)

        # Copiando FTF
        pyautogui.click(1086,649)
        pyautogui.click(1086,649)
        pyautogui.click(button='RIGHT')
        time.sleep(0.2)
        for _ in range(3):
            pyautogui.press('down')
        pyautogui.press('ENTER')
        FTF = pyperclip.paste()
        time.sleep(0.1)
        print(FTF)
        
        ancora_break = FTF
        # Clicando em Cancelar
        pyautogui.press('TAB')
        pyautogui.press('TAB')
        pyautogui.press('ENTER')
        time.sleep(0.30)
        
        # Inserindo dados na planilha
        variaveis = [Fabricante, Num_Rolamento, BPFO, BPFI, BSF, FTF]

        for i, variavel in enumerate(variaveis, start=1):
            sheet.cell(row=linha, column=i, value=variavel)

        linha += 1 
        rolamento += 1
